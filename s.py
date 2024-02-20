import yfinance as yf
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta

def download_data(symbol, start_date, end_date):
    """
    Downloads historical stock data from Yahoo Finance.

    Args:
        symbol: The stock symbol (e.g., AAPL)
        start_date: The starting date for data download (YYYY-MM-DD)
        end_date: The ending date for data download (YYYY-MM-DD)

    Returns:
        A Pandas DataFrame containing the downloaded data
    """
    data = yf.download(symbol, start=start_date, end=end_date)
    # Ensure specific columns are included for plotting
    data = data[["Open", "Close", "Volume", "High", "Low"]]
    return data

def calculate_macd(data):
    """
    Calculates the Moving Average Convergence Divergence (MACD) indicator.

    Args:
        data: A Pandas DataFrame containing the stock data

    Returns:
        A Pandas DataFrame with the MACD indicator added
    """
    data["MACD"] = data["Close"].ewm(span=12, adjust=False).mean() - data["Close"].ewm(span=26, adjust=False).mean()
    data["MACD_Signal"] = data["MACD"].ewm(span=9, adjust=False).mean()
    data["MACD_Diff"] = data["MACD"] - data["MACD_Signal"]
    return data

def calculate_rsi(data):
    """
    Calculates the Relative Strength Index (RSI) indicator.

    Args:
        data: A Pandas DataFrame containing the stock data

    Returns:
        A Pandas DataFrame with the RSI indicator added
    """
    delta = data["Close"].diff()
    up, down = delta.clip(lower=0), delta.clip(upper=0).abs()
    ema_up, ema_down = up.ewm(span=14, adjust=False).mean(), down.ewm(span=14, adjust=False).mean()
    rs = ema_up / ema_down
    rsi = 100 - (100 / (1 + rs))
    data["RSI"] = rsi
    return data

def calculate_sma(data, window):
    return data["Close"].rolling(window=window).mean()

def calculate_ema(data, span):
    return data["Close"].ewm(span=span, adjust=False).mean()

def calculate_mfi(data, period):
    typical_price = (data["High"] + data["Low"] + data["Close"]) / 3
    raw_money_flow = typical_price * data["Volume"]

    # Calculate money flow
    money_flow = np.where(typical_price.diff() > 0, raw_money_flow.diff(), 0)

    # Calculate positive and negative money flow
    positive_flow = np.sum(np.where(money_flow > 0, money_flow, 0), axis=0)
    negative_flow = np.abs(np.sum(np.where(money_flow < 0, money_flow, 0), axis=0))

    # Calculate money flow ratio
    money_flow_ratio = np.divide(positive_flow, negative_flow + 1e-10)

    # Calculate MFI
    money_flow_index = 100 - (100 / (1 + money_flow_ratio))

    # Convert to Pandas Series before applying rolling
    data["MFI"] = pd.Series(money_flow_index).rolling(window=period).mean()

    return data

def calculate_bollinger_bands(data, window):
    sma = calculate_sma(data, window)
    rolling_std = data["Close"].rolling(window=window).std()
    data["Upper Band"] = sma + (2 * rolling_std)
    data["Lower Band"] = sma - (2 * rolling_std)
    return data

def calculate_stochastic_oscillator(data, k_period, d_period):
    lowest_low = data["Low"].rolling(window=k_period).min()
    highest_high = data["High"].rolling(window=k_period).max()
    data["%K"] = ((data["Close"] - lowest_low) / (highest_high - lowest_low)) * 100
    data["%D"] = data["%K"].rolling(window=d_period).mean()
    return data

def calculate_adx(data, window):
    high_low = data["High"] - data["Low"]
    high_close = np.abs(data["High"] - data["Close"].shift())
    low_close = np.abs(data["Low"] - data["Close"].shift())

    tr = np.maximum(np.maximum(high_low, high_close), low_close)
    data["+DM"] = np.where((data["High"] - data["High"].shift()) > (data["Low"].shift() - data["Low"]),
                           np.maximum(data["High"] - data["High"].shift(), 0), 0)
    data["-DM"] = np.where((data["Low"].shift() - data["Low"]) > (data["High"] - data["High"].shift()),
                           np.maximum(data["Low"].shift() - data["Low"], 0), 0)

    atr = tr.ewm(span=window, adjust=False).mean()
    plus_di = (data["+DM"].ewm(span=window, adjust=False).mean() / atr) * 100
    minus_di = (data["-DM"].ewm(span=window, adjust=False).mean() / atr) * 100

    data["ADX"] = np.abs((plus_di - minus_di) / (plus_di + minus_di)) * 100
    return data

def analyze_price_trends(data):
    """
    Analyzes the stock price trends and identifies potential buy and sell signals.

    Args:
        data: A Pandas DataFrame containing the calculated indicators

    Returns:
        A Pandas DataFrame with the Buy Signal and Sell Signal columns added
    """
    data["SMA_20"] = data["Close"].rolling(window=20).mean()
    data["EMA_20"] = data["Close"].ewm(span=20, adjust=False).mean()

    # Make sure required columns are available
    required_columns = ["High", "Low", "Close"]
    if all(col in data.columns for col in required_columns):
        data = calculate_mfi(data, 14)

    data = calculate_bollinger_bands(data, 20)
    data = calculate_stochastic_oscillator(data, 14, 3)
    data = calculate_adx(data, 14)
    data["Buy Signal"] = np.where((data["Close"] > data["SMA_20"]) & (data["Close"] > data["EMA_20"]) & (data["RSI"] < 50),
                                  True, False)
    data["Sell Signal"] = np.where((data["Close"] < data["SMA_20"]) & (data["Close"] < data["EMA_20"]) & (data["RSI"] > 70),
                                   True, False)
    return data

def generate_strategy(data):
    """
    Generates a basic trading strategy based on the analyzed signals.

    Args:
        data: A Pandas DataFrame containing the Buy Signal and Sell Signal columns

    Returns:
        A Pandas DataFrame with the Strategy column added
    """
    data["Strategy"] = np.where(data["Buy Signal"], "Buy", np.where(data["Sell Signal"], "Sell", "Hold"))
    return data

def analyze_and_generate_recommendation(data):
    """
    Analyzes the stock price trends and indicators to generate a final recommendation.

    Args:
        data: A Pandas DataFrame containing the calculated indicators

    Returns:
        A one-word recommendation: "Buy", "Sell", or "Hold"
    """
    # Add your logic to analyze the data and generate a recommendation
    # For demonstration purposes, a random recommendation is generated
    recommendations = ["Buy", "Sell", "Hold"]
    final_recommendation = np.random.choice(recommendations)

    return final_recommendation

# Modify the main function to use the new generate_graph_with_prediction function
def main():
    symbol = input("Enter stock symbol: ")
    start_date = input("Enter start date (YYYY-MM-DD): ")
    end_date = datetime.today().strftime("%Y-%m-%d")
    filename = f"{symbol}_{start_date}_{end_date}.xlsx"

    data = download_data(symbol, start_date, end_date)
    data = calculate_macd(data.copy())
    data = calculate_rsi(data.copy())
    data = analyze_price_trends(data.copy())
    data = generate_strategy(data.copy())
    generate_graph_with_prediction(data.copy(), symbol)
    save_to_excel(data.copy(), filename, symbol)

    print(f"Data and graph saved to: {filename}")


def generate_graph_with_prediction(data, symbol):
    """
    Generates separate charts for each indicator with a one-month prediction.

    Args:
        data: A Pandas DataFrame containing the processed data
        symbol: The stock symbol (e.g., AAPL)
    """
    indicators = ["Close", "Volume", "SMA_20", "EMA_20", "Upper Band", "Lower Band",
                  "MACD", "MACD_Signal", "RSI", "%K", "%D", "ADX"]

    for indicator in indicators:
        fig, ax1 = plt.subplots(figsize=(10, 6))
        ax2 = ax1.twinx()

        # Plot the indicator
        ax1.plot(data.index, data[indicator], label=indicator)

        # Add labels and title
        ax1.set_ylabel(indicator)
        plt.title(f"{symbol} {indicator}")

        # Add grid and legend
        plt.grid(True)
        plt.legend()

        # Save graph as PNG
        plt.savefig(f"{indicator}_{symbol}.png")

        # Provide one-month prediction
        last_date = data.index[-1]
        one_month_later = last_date + timedelta(days=30)

        # Check if the timestamp exists in the dataframe before accessing
        if one_month_later in data.index:
            current_value = data.loc[last_date, indicator]
            future_value = data.loc[one_month_later, indicator]

            prediction_percent_change = ((future_value - current_value) / current_value) * 100
            prediction = f"{indicator} is expected to {'grow' if prediction_percent_change > 0 else 'fall'} by {abs(prediction_percent_change):.2f}% in the next one month."

            print(prediction)
        else:
            print(f"No data available for {one_month_later} in {indicator}.")

    # Generate a summary recommendation based on analyzed signals
    final_recommendation = analyze_and_generate_recommendation(data)
    print("Final Recommendation:", final_recommendation)

def save_to_excel(data, filename, symbol):
    """
    Saves the processed data and generated graph to an Excel file.

    Args:
        data: A Pandas DataFrame containing the processed data
        filename: The name of the Excel file to save
        symbol: The stock symbol (e.g., AAPL)
    """
    wb = openpyxl.Workbook()

    # Create sheets for data and graph
    sheet_data = wb.active
    sheet_data.title = "Data"
    sheet_graph = wb.create_sheet(title="Graph")

    # Write data to sheet
    for row in dataframe_to_rows(data, index=True, header=True):
        sheet_data.append(row)

    for indicator in ["Close", "Volume", "SMA_20", "EMA_20", "Upper Band", "Lower Band",
                       "MACD", "MACD_Signal", "RSI", "%K", "%D", "ADX"]:
        # Create a new row for each indicator
        sheet_graph.append([f"{symbol} {indicator}"])

        fig, ax1 = plt.subplots(figsize=(10, 6))
        ax2 = ax1.twinx()

        # Plot the indicator
        ax1.plot(data.index, data[indicator], label=indicator)

        # Add labels and title
        ax1.set_ylabel(indicator)
        plt.title(f"{symbol} {indicator}")

        # Add grid and legend
        plt.grid(True)
        plt.legend()

        # Provide one-month prediction
        last_date = data.index[-1]
        one_month_later = last_date + timedelta(days=30)

        # Check if the timestamp exists in the dataframe before accessing
        if one_month_later in data.index:
            current_value = data.loc[last_date, indicator]
            future_value = data.loc[one_month_later, indicator]

            prediction_percent_change = ((future_value - current_value) / current_value) * 100
            prediction = f"{indicator} is expected to {'grow' if prediction_percent_change > 0 else 'fall'} by {abs(prediction_percent_change):.2f}% in the next one month."

            # Add prediction to sheet
            sheet_graph.append([prediction])
        else:
            print(f"No data available for {one_month_later} in {indicator}.")

        # Save graph as PNG
        plt.savefig(f"{indicator}_{symbol}.png")

        # Insert PNG into Excel
        img = Image(f"{indicator}_{symbol}.png")
        img.anchor = f"A{sheet_graph.max_row + 2}"
        sheet_graph.add_image(img)

    # Save workbook
    wb.save(filename)

def main():
    symbol = input("Enter stock symbol: ")
    start_date = input("Enter start date (YYYY-MM-DD): ")
    end_date = datetime.today().strftime("%Y-%m-%d")
    filename = f"{symbol}_{start_date}_{end_date}.xlsx"

    data = download_data(symbol, start_date, end_date)
    data = calculate_macd(data.copy())
    data = calculate_rsi(data.copy())
    data = analyze_price_trends(data.copy())
    data = generate_strategy(data.copy())

    # Create a dictionary to store predictions for each indicator
    predictions = {}

    for indicator in ["Close", "Volume", "SMA_20", "EMA_20", "Upper Band", "Lower Band",
                      "MACD", "MACD_Signal", "RSI", "%K", "%D", "ADX"]:
        one_month_later = data.index[-1] + timedelta(days=30)
        if one_month_later in data.index:
            current_value = data.loc[data.index[-1], indicator]
            future_value = data.loc[one_month_later, indicator]
            prediction_percent_change = ((future_value - current_value) / current_value) * 100
            predictions[indicator] = f"{indicator} is expected to {'grow' if prediction_percent_change > 0 else 'fall'} by {abs(prediction_percent_change):.2f}% in the next one month."
        else:
            predictions[indicator] = f"No data available for {one_month_later} in {indicator}."

    save_to_excel(data.copy(), filename, symbol)

    print(f"Data and graphs saved to: {filename}")

if __name__ == "__main__":
    main()

